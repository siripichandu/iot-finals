import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

LOG_FILE = "logs/subscriber.log"
OUT_FILE = "data/sensor_data.xlsx"

# handles BOTH formats:
# old: [KITCHEN] Temp=21.2°C  Humid=35.1%  AQI=N/A  Motion=0
# new: [BEDROOM ] Temp= 19.8°C  Humid= 38.1%  AQI=  81  Motion=0  TOD=afternoon
pattern = re.compile(
    r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}).*?"
    r"\[INFO\]\s+\[(\w+)\s*\].*?"
    r"Temp=\s*([\d.]+).*?"
    r"Humid=\s*([\d.]+).*?"
    r"AQI=\s*([\d.]+|N/A).*?"
    r"Motion=(\d)"
)

records = []
with open(LOG_FILE, "r") as f:
    for line in f:
        m = pattern.search(line)
        if m:
            aqi_raw = m.group(5).strip()
            records.append({
                "Timestamp":   m.group(1),
                "Room":        m.group(2).strip().title(),
                "Temperature": float(m.group(3)),
                "Humidity":    float(m.group(4)),
                "AQI":         None if aqi_raw == "N/A" else float(aqi_raw),
                "Motion":      int(m.group(6)),
            })

df = pd.DataFrame(records)
df["Timestamp"] = pd.to_datetime(df["Timestamp"])
df = df.sort_values("Timestamp").reset_index(drop=True)

def aqi_label(v):
    if v is None or (isinstance(v, float) and pd.isna(v)): return "N/A"
    if v < 300:  return "Good"
    if v < 800:  return "Moderate"
    if v < 1500: return "Poor"
    return "Hazardous"

df["AQI_Label"]    = df["AQI"].apply(aqi_label)
df["Motion_Label"] = df["Motion"].map({0: "Empty", 1: "Occupied"})

summary_rows = []
for room, grp in df.groupby("Room"):
    aqi_clean = grp["AQI"].dropna()
    summary_rows.append({
        "Room":             room,
        "Total Records":    len(grp),
        "Avg Temp (°C)":    round(grp["Temperature"].mean(), 2),
        "Min Temp (°C)":    round(grp["Temperature"].min(), 2),
        "Max Temp (°C)":    round(grp["Temperature"].max(), 2),
        "Avg Humidity (%)": round(grp["Humidity"].mean(), 2),
        "Min Humidity (%)": round(grp["Humidity"].min(), 2),
        "Max Humidity (%)": round(grp["Humidity"].max(), 2),
        "Avg AQI":          round(aqi_clean.mean(), 2) if len(aqi_clean) else "N/A",
        "Max AQI":          round(aqi_clean.max(), 2)  if len(aqi_clean) else "N/A",
        "Motion Events":    int(grp["Motion"].sum()),
        "First Reading":    grp["Timestamp"].min().strftime("%Y-%m-%d %H:%M"),
        "Last Reading":     grp["Timestamp"].max().strftime("%Y-%m-%d %H:%M"),
    })
summary_df = pd.DataFrame(summary_rows)

with pd.ExcelWriter(OUT_FILE, engine="openpyxl", datetime_format="YYYY-MM-DD HH:MM:SS") as writer:
    df.to_excel(writer, sheet_name="All Data", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    for room, grp in df.groupby("Room"):
        grp.reset_index(drop=True).to_excel(writer, sheet_name=room, index=False)

wb = load_workbook(OUT_FILE)

HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Arial")
ALT_FILL    = PatternFill("solid", fgColor="F4F6FA")
BORDER      = Border(bottom=Side(style="thin", color="DDDDDD"))

AQI_FILLS = {
    "Good":      PatternFill("solid", fgColor="C6EFCE"),
    "Moderate":  PatternFill("solid", fgColor="FFEB9C"),
    "Poor":      PatternFill("solid", fgColor="FFCC99"),
    "Hazardous": PatternFill("solid", fgColor="FFC7CE"),
}
MOTION_FILLS = {
    "Occupied": PatternFill("solid", fgColor="FCE4EC"),
    "Empty":    PatternFill("solid", fgColor="E8F5E9"),
}

def style_sheet(ws):
    headers = {cell.value: cell.column for cell in ws[1]}
    aqi_col    = headers.get("AQI_Label")
    motion_col = headers.get("Motion_Label")

    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            if i % 2 == 0:
                cell.fill = ALT_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border    = BORDER
            if cell.column == 1:
                cell.number_format = "YYYY-MM-DD HH:MM:SS"
                cell.alignment = Alignment(horizontal="left")

        if aqi_col:
            c = ws.cell(row=i, column=aqi_col)
            if c.value in AQI_FILLS:
                c.fill = AQI_FILLS[c.value]
                c.font = Font(bold=True, name="Arial")

        if motion_col:
            c = ws.cell(row=i, column=motion_col)
            if c.value in MOTION_FILLS:
                c.fill = MOTION_FILLS[c.value]

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

for name in wb.sheetnames:
    style_sheet(wb[name])

wb.save(OUT_FILE)
print(f"Done — {len(df)} records | Rooms: {df['Room'].value_counts().to_dict()}")
print(f"Saved: {OUT_FILE}")